from collections import defaultdict

from django.shortcuts import render, redirect
from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework.authtoken.models import Token
from rest_framework.views import APIView
from rest_framework.authentication import TokenAuthentication
from django.contrib.auth import authenticate
from django.db.models import Prefetch, Sum
from django.db import IntegrityError
from django.http import HttpResponse
import pandas as pd
import random
import os
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Image, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from django.conf import settings

# Import your models and serializers
from .models import (
    Gender, Discipline, School,
    Competitor, Season, AgeGroup,
    Stage, CompetitionDay, Registration,
    Result
)
from .serializers import (
    TokenSerializer, SchoolSerializer, GenderSerializer,
    DisciplineSerializer, CompetitorSerializer,
    SeasonSerializer, AgeGroupSerializer, StageSerializer,
    CompetitionDaySerializer, RegistrationSerializer,
    ResultSerializer, RandomizeBibNumbersSerializer
)


class ObtainAuthTokenView(APIView):
    throttle_classes = ()
    permission_classes = ()
    authentication_classes = ()

    @swagger_auto_schema(
        operation_description="Post username and password to get an authentication token",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['username', 'password'],
            properties={
                'username': openapi.Schema(type=openapi.TYPE_STRING, description='Username'),
                'password': openapi.Schema(type=openapi.TYPE_STRING, description='Password'),
            },
        ),
        responses={
            200: TokenSerializer,
            400: 'Invalid username/password'
        }
    )
    def post(self, request, *args, **kwargs):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            token, created = Token.objects.get_or_create(user=user)
            return Response({'token': token.key})
        else:
            return Response({"error": "Wrong Credentials"}, status=status.HTTP_400_BAD_REQUEST)


class LogoutView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_description="Logout a user by invalidating the authentication token",
        responses={
            200: 'Successfully logged out',
            401: 'Unauthorized'
        }
    )
    def post(self, request, *args, **kwargs):
        request.user.auth_token.delete()
        return Response({"message": "Successfully logged out"}, status=status.HTTP_200_OK)


class SchoolListCreate(generics.ListCreateAPIView):
    queryset = School.objects.all()
    serializer_class = SchoolSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class SchoolRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = School.objects.all()
    serializer_class = SchoolSerializer
    permission_classes = [IsAuthenticated]


class DisciplineListCreate(generics.ListCreateAPIView):
    queryset = Discipline.objects.all()
    serializer_class = DisciplineSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class DisciplineRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Discipline.objects.all()
    serializer_class = DisciplineSerializer
    permission_classes = [IsAuthenticated]


class GenderListCreate(generics.ListCreateAPIView):
    queryset = Gender.objects.all()
    serializer_class = GenderSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class GenderRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Gender.objects.all()
    serializer_class = GenderSerializer
    permission_classes = [IsAuthenticated]


class CompetitorListCreate(generics.ListCreateAPIView):
    queryset = Competitor.objects.all()
    serializer_class = CompetitorSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class CompetitorRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Competitor.objects.all()
    serializer_class = CompetitorSerializer
    permission_classes = [IsAuthenticated]


class SeasonListCreate(generics.ListCreateAPIView):
    queryset = Season.objects.all()
    serializer_class = SeasonSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class SeasonRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Season.objects.all()
    serializer_class = SeasonSerializer
    permission_classes = [IsAuthenticated]


class AgeGroupListCreate(generics.ListCreateAPIView):
    queryset = AgeGroup.objects.all()
    serializer_class = AgeGroupSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class AgeGroupRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = AgeGroup.objects.all()
    serializer_class = AgeGroupSerializer
    permission_classes = [IsAuthenticated]


class StageListCreate(generics.ListCreateAPIView):
    queryset = Stage.objects.all()
    serializer_class = StageSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class StageRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Stage.objects.all()
    serializer_class = StageSerializer
    permission_classes = [IsAuthenticated]


class CompetitionDayListCreate(generics.ListCreateAPIView):
    queryset = CompetitionDay.objects.all()
    serializer_class = CompetitionDaySerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class CompetitionDayRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = CompetitionDay.objects.all()
    serializer_class = CompetitionDaySerializer
    permission_classes = [IsAuthenticated]


class RegistrationListCreate(generics.ListCreateAPIView):
    serializer_class = RegistrationSerializer
    # If you need different permissions for GET vs POST:
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    def get_queryset(self):
        """
        Optionally filter by ?date=YYYY-MM-DD
        """
        queryset = Registration.objects.all()
        date_str = self.request.query_params.get('date')

        if date_str:
            # Filter by competition_day.date
            queryset = queryset.filter(competition_day__date=date_str)
            # e.g. /api/registration/?date=2024-04-13

        return queryset


class RegistrationRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Registration.objects.all()
    serializer_class = RegistrationSerializer
    permission_classes = [IsAuthenticated]


class ResultListCreate(generics.ListCreateAPIView):
    queryset = Result.objects.all()
    serializer_class = ResultSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class ResultRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Result.objects.all()
    serializer_class = ResultSerializer
    permission_classes = [IsAuthenticated]

    def update(self, request, *args, **kwargs):
        # Force partial update logic if you want a full PUT to behave like PATCH
        kwargs['partial'] = True
        return super().update(request, *args, **kwargs)



class SearchResultsAPIView(generics.ListAPIView):
    """
    Example search:
      GET /api/results/search?season=2024-2025&stage=Stage+1&discipline=Slalom
    """
    serializer_class = ResultSerializer

    def get_queryset(self):
        queryset = Result.objects.all()
        season = self.request.query_params.get('season')
        stage = self.request.query_params.get('stage')
        discipline = self.request.query_params.get('discipline')

        if season and stage and discipline:
            queryset = queryset.filter(
                registration__competition_day__stage__season__name=season,
                registration__competition_day__stage__name=stage,
                registration__competition_day__discipline__name=discipline
            )
        return queryset



@swagger_auto_schema(method='post', request_body=RandomizeBibNumbersSerializer)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def randomize_bib_numbers(request):
    """
    POST body example:
        {
          "start_number": 1,
          "ignore_numbers": [],
          "competition_day_id": 7,
          "age_group_id": null
        }

    We do:
      - Split regs into female vs. male by `age_group.gender`.
      - Sort each dictionary's age groups from youngest -> oldest.
      - Assign bibs to female groups first, from `start_number`.
      - Then assign bibs to male groups, also from `start_number`.
      - Each dictionary is truly separate, so no mixing of female & male age groups.
    """
    serializer = RandomizeBibNumbersSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)

    start_number = serializer.validated_data.get('start_number', 1)
    ignore_numbers = set(serializer.validated_data.get('ignore_numbers', []))
    competition_day_id = serializer.validated_data.get('competition_day_id')
    age_group_id = serializer.validated_data.get('age_group_id', None)

    # Filter registrations
    filters = {"competition_day_id": competition_day_id}
    if age_group_id:
        filters["age_group_id"] = age_group_id

    registrations = (
        Registration.objects
        .filter(**filters)
        .select_related('age_group', 'competitor')
    )
    if not registrations.exists():
        return Response({"message": "No registrations found."}, status=200)

    # Split into female vs. male
    female_regs = []
    male_regs = []

    for reg in registrations:
        # Carefully handle the 'gender' to avoid trailing spaces
        ag_gender = (reg.age_group.gender.name or "").strip() if reg.age_group else ""
        if ag_gender == 'ქალი':
            female_regs.append(reg)
        elif ag_gender == 'კაცი':
            male_regs.append(reg)
        else:
            # If there's unknown gender data, you can either skip or handle it
            # For demonstration, we'll skip them
            pass

    # Group each list by age_group
    def group_by_age_group(reg_list):
        d = defaultdict(list)
        for r in reg_list:
            d[r.age_group].append(r)
        return d

    female_dict = group_by_age_group(female_regs)  # purely female
    male_dict = group_by_age_group(male_regs)      # purely male

    # Sorting key to go from "youngest => oldest"
    # birth_year_start: bigger => younger => sort descending
    def age_group_sort_key(ag):
        # If birth_year_start is None => treat as 0 => older or open => last
        val = ag.birth_year_start or 0
        # negative => largest => first
        return -val

    def randomize_one_gender(group_dict, initial_bib, ignore_nums):
        """
        Sort each age group from youngest->oldest,
        shuffle the registrations,
        assign consecutive bibs skipping ignore_nums.
        """
        current_bib = initial_bib
        # Sort age_group keys
        sorted_ags = sorted(group_dict.keys(), key=age_group_sort_key)

        for ag in sorted_ags:
            regs_in_group = group_dict[ag]
            if not regs_in_group:
                continue

            random.shuffle(regs_in_group)
            needed = len(regs_in_group)

            assigned_bibs = []
            candidate = current_bib
            while len(assigned_bibs) < needed:
                if candidate not in ignore_nums:
                    assigned_bibs.append(candidate)
                candidate += 1

            for i, reg in enumerate(regs_in_group):
                reg.bib_number = assigned_bibs[i]
                reg.save()

            current_bib = candidate

        return current_bib

    # 1) Randomize female
    randomize_one_gender(female_dict, start_number, ignore_numbers)

    # 2) Randomize male, from start_number again
    randomize_one_gender(male_dict, start_number, ignore_numbers)

    return Response({"message": "Bib numbers randomized successfully"}, status=200)




def create_result_for_registration(registration_id):
    try:
        registration = Registration.objects.get(pk=registration_id)
        # Check if a Result already exists for this registration
        if not Result.objects.filter(registration=registration).exists():
            # Create a new Result
            Result.objects.create(
                registration=registration
                # You could set default times or place=..., etc., if needed
            )
        else:
            pass  # Already has a result
    except Registration.DoesNotExist:
        return Response({"error": f"Registration with ID {registration_id} does not exist"},
                        status=status.HTTP_404_NOT_FOUND)
    except IntegrityError:
        return Response({"error": f"Error creating result for Registration ID {registration_id}"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response({"message": f"Result added successfully for Registration ID {registration_id}"},
                    status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def batch_sync_results(request):
    registration_ids = request.data.get('registration_ids', [])
    for reg_id in registration_ids:
        create_result_for_registration(reg_id)
    return Response({"message": "Results added successfully"}, status=status.HTTP_200_OK)


class SyncRegistrationToResultsView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_description="Synchronize Registration entries to Result table",
        responses={
            200: 'Synchronization completed successfully',
            500: 'Internal Server Error'
        }
    )
    def post(self, request, *args, **kwargs):
        try:
            sync_registration_to_results()
            return Response({"message": "Synchronization completed successfully"}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def sync_registration_to_results():
    for reg in Registration.objects.all():
        if not Result.objects.filter(registration=reg).exists():
            try:
                Result.objects.create(registration=reg)
            except IntegrityError:
                pass  # or handle it

# views.py
@api_view(['GET'])
@permission_classes([AllowAny])
def results_by_date(request):
    """
    GET /api/results?date=YYYY-MM-DD
    Returns all 'Results' for that date's competition day(s).
    """
    date_str = request.GET.get('date', None)
    if not date_str:
        return Response({"error": "No date specified"}, status=400)

    try:
        # Attempt to parse date
        from datetime import datetime
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return Response({"error": "Invalid date format, expected YYYY-MM-DD"}, status=400)

    # Filter all Results whose competition_day.date == date_obj
    results = Result.objects.filter(registration_id__competition_day_id__date=date_obj)
      # or whatever your model field is: e.g. competition_day__date=date_obj
      # if your model is 'competition_day', do `competition_day__date=date_obj`

    serializer = ResultSerializer(results, many=True)
    return Response(serializer.data, status=200)




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def download_excel(request):
    if request.method == 'POST':
        registration_ids = request.data.get('registrationIds', [])

        # Fetch registrations
        registrations = (
            Registration.objects
            .filter(id__in=registration_ids)
            .select_related('competitor', 'competition_day', 'age_group')
        )

        if not registrations.exists():
            return Response({"message": "No registrations found."}, status=200)

        # Build a nested structure: { comp_day -> { age_group -> [regs] } }
        from collections import defaultdict
        grouped_by_day = defaultdict(lambda: defaultdict(list))

        for reg in registrations:
            grouped_by_day[reg.competition_day][reg.age_group].append(reg)

        # Prepare style helpers for Excel
        bold_font = Font(bold=True)
        sky_blue_fill = PatternFill(
            start_color='00B0F0',
            end_color='00B0F0',
            fill_type='solid'
        )

        # We'll keep a list of DataFrames
        dfs = []


        # For each day, we create a row for the day, then sub-rows for each age group
        for comp_day, ag_map in grouped_by_day.items():
            # This day header
            day_header_text = f"{comp_day.discipline.name} on {comp_day.date}"
            day_header_df = pd.DataFrame([{'BIB': 'Competition Day:', 'Name': day_header_text}])
            dfs.append(day_header_df)

            # Now, for each age_group, we produce a sub-header and data
            # Sort AgeGroups if you want a specific order (e.g., youngest->oldest)
            # For example, descending birth_year_start
            def ag_sort_key(ag):
                """
                1) If ag.gender == 'ქალი', gender_priority = 0
                   If ag.gender == 'კაცი', gender_priority = 1
                2) birth_year_start = ag.birth_year_start or 0
                   Negative => descending
                """
                gender_priority = 0 if ag.gender.name == 'ქალი' else 1
                start_val = ag.birth_year_start or 0
                return (gender_priority, -start_val)

            sorted_ags = sorted(ag_map.keys(), key=ag_sort_key)

            for ag in sorted_ags:
                regs_in_ag = ag_map[ag]
                # Sort by bib ascending
                regs_sorted = sorted(regs_in_ag, key=lambda r: (r.bib_number if r.bib_number else 999999))

                # Make a sub-header row for the age group
                # E.g.  "ქალი 2010-2011" or "ქალი None-2007"
                if ag.birth_year_start is None:
                    year_range = f"↓ - {ag.birth_year_end}" if ag.birth_year_end else "↓"
                elif ag.birth_year_end:
                    year_range = f"{ag.birth_year_start}-{ag.birth_year_end}"
                else:
                    year_range = f"{ag.birth_year_start}+"

                ag_header_text = f"{ag.gender} {year_range}"
                ag_header_df = pd.DataFrame([{'BIB': 'AgeGroup:', 'Name': ag_header_text}])
                dfs.append(ag_header_df)

                # Now the data rows
                rows = []
                for r in regs_sorted:
                    c = r.competitor
                    rows.append({
                        'BIB': r.bib_number,
                        'Name': f"{c.first_name} {c.last_name}",
                        'Gender': c.gender.name if c.gender else '',  # or c.gender if it's a string
                        'Year': c.year_of_birth,
                        'School': c.school.name if c.school else ''
                    })
                df = pd.DataFrame(rows)
                dfs.append(df)

        # Concatenate all DataFrames
        all_data = pd.concat(dfs, ignore_index=True)

        # Build response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Registrations.xlsx"'

        # Write to Excel
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            all_data.to_excel(writer, sheet_name='Registrations', index=False)
            worksheet = writer.sheets['Registrations']

            # Apply bold font and sky blue background to day headers and AgeGroup rows
            # Those rows where the first cell is 'Competition Day:' or 'AgeGroup:'
            for row in worksheet.iter_rows(min_row=1, max_col=2, max_row=worksheet.max_row):
                first_cell = row[0].value
                if first_cell in ('Competition Day:', 'AgeGroup:'):
                    for cell in row:
                        cell.font = bold_font
                        cell.fill = sky_blue_fill

        return response

    return Response({"error": "Invalid request"}, status=400)





@api_view(['POST'])
@permission_classes([IsAuthenticated])
def download_pdf(request):
    if request.method == 'POST':
        registration_ids = request.data.get('registrationIds', [])

        font_path = os.path.join(settings.BASE_DIR, 'competition/fonts', 'bpg_glaho_sylfaen.ttf')
        try:
            pdfmetrics.registerFont(TTFont('kartuli', font_path))
        except Exception as e:
            return Response({"error": f"Failed to register font: {str(e)}"}, status=500)

        # Fetch registrations
        registrations = (
            Registration.objects
            .filter(id__in=registration_ids)
            .select_related('competitor', 'competition_day', 'age_group')
        )
        if not registrations.exists():
            return Response({"message": "No registrations found."}, status=200)

        from collections import defaultdict
        grouped_by_day = defaultdict(lambda: defaultdict(list))

        for reg in registrations:
            grouped_by_day[reg.competition_day][reg.age_group].append(reg)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Start_list.pdf"'
        doc = SimpleDocTemplate(response, pagesize=letter, topMargin=1.5 * inch)
        elements = []

        sky_color = colors.HexColor('#2F89D0')
        sand_color = colors.HexColor('#b0caff')
        page_width, page_height = letter
        table_width = page_width * 0.9
        # We'll define columns as [BIB, Name, Gender, YOB, School]
        column_widths = [
            table_width / 12,
            (table_width - (table_width / 12) * 3) / 2,  # wide
            table_width / 12,  # Gender
            table_width / 12,  # YOB
            (table_width - (table_width / 12) * 3) / 2,  # School
        ]

        def draw_header(canvas, doc):
            logo_path = os.path.join(settings.BASE_DIR, 'competition/fonts', 'GSF-Logo.png')
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=80, height=80)
                logo.drawOn(canvas, doc.leftMargin - 40, doc.height + doc.topMargin - 30)

            title = "Start List"
            title_style = ParagraphStyle(name='TitleStyle', fontSize=16, alignment=1)
            p_title = Paragraph(title, title_style)
            p_title.wrapOn(canvas, doc.width, doc.topMargin)
            p_title.drawOn(canvas, doc.width / 2 - 160, doc.height + doc.topMargin)

            from datetime import datetime
            date_text = "Date: " + str(datetime.now().strftime('%Y-%m-%d'))
            p_date = Paragraph(date_text, title_style)
            p_date.wrapOn(canvas, doc.width, doc.topMargin)
            p_date.drawOn(canvas, doc.width - doc.rightMargin - 110, doc.height + doc.topMargin)

        # We'll define a sort for AgeGroups if you want the same youngest->oldest logic
        def ag_sort_key(ag):
            """
            1) If ag.gender == 'ქალი', gender_priority = 0
               If ag.gender == 'კაცი', gender_priority = 1
            2) birth_year_start = ag.birth_year_start or 0
               Negative => descending
            """
            gender_priority = 0 if ag.gender.name == 'ქალი' else 1
            start_val = ag.birth_year_start or 0
            return (gender_priority, -start_val)

        for comp_day, ag_map in grouped_by_day.items():
            # Day header
            day_header_data = [[f"{comp_day.discipline.name} on {comp_day.date}"]]
            day_header_table = Table(day_header_data, colWidths=[table_width])
            day_header_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), sky_color),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONT', (0, 0), (-1, -1), 'kartuli', 12),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ])
            day_header_table.setStyle(day_header_style)
            elements.append(day_header_table)
            elements.append(Spacer(1, 12))

            # Now for each age_group in sorted order
            sorted_ags = sorted(ag_map.keys(), key=ag_sort_key)

            for ag in sorted_ags:
                regs_in_ag = ag_map[ag]
                if not regs_in_ag:
                    continue

                # Sort by bib ascending
                regs_sorted = sorted(regs_in_ag, key=lambda r: (r.bib_number or 999999))

                # AgeGroup header row
                if not ag.birth_year_start:
                    year_range = f"↓ - {ag.birth_year_end}" if ag.birth_year_end else "↓"
                elif ag.birth_year_end:
                    year_range = f"{ag.birth_year_start}-{ag.birth_year_end}"
                else:
                    year_range = f"{ag.birth_year_start}+"

                ag_header_data = [[f"{ag.gender} {year_range}"]]
                ag_header_table = Table(ag_header_data, colWidths=[table_width])
                ag_header_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), sand_color),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONT', (0, 0), (-1, -1), 'kartuli', 10),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ])
                ag_header_table.setStyle(ag_header_style)
                elements.append(ag_header_table)
                elements.append(Spacer(1, 12))

                # Build table data
                data = [['BIB', 'სპორტსმენი', 'სქესი', 'წელი', 'სკოლა']]
                for r in regs_sorted:
                    c = r.competitor
                    data.append([
                        r.bib_number if r.bib_number else '',
                        f"{c.first_name} {c.last_name}",
                        c.gender.name if c.gender else '',
                        c.year_of_birth,
                        c.school.name if c.school else ''
                    ])

                data_table = Table(data, colWidths=column_widths)
                data_style = TableStyle([
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONT', (0, 0), (-1, -1), 'kartuli', 10),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                ])
                data_table.setStyle(data_style)
                elements.append(data_table)
                elements.append(Spacer(1, 12))

        doc.build(elements, onFirstPage=draw_header)
        return response

    return Response({"error": "Invalid request"}, status=400)




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def download_results_pdf(request):
    """
    POST body:
      {
        "resultIds": [ 12, 55, 41, ... ]  # in EXACT order from the front end
      }
    We'll group results by:
      - result.registration.competition_day
      - result.registration.age_group
    BUT we preserve the front-end's EXACT order in each day/age group
    (no re-sorting).
    """
    if request.method != 'POST':
        return Response({"error": "Invalid request"}, status=400)

    result_ids = request.data.get('resultIds', [])
    if not result_ids:
        return Response({"message": "No result IDs provided."}, status=400)

    # 1) Register font if needed
    font_path = os.path.join(settings.BASE_DIR, 'competition/fonts', 'bpg_glaho_sylfaen.ttf')
    try:
        pdfmetrics.registerFont(TTFont('kartuli', font_path))
    except:
        pass

    # 2) Build an index_map so we can preserve the EXACT order from the front-end
    index_map = {rid: i for i, rid in enumerate(result_ids)}

    # 3) Fetch all Results
    results_qs = (
        Result.objects
        .filter(id__in=result_ids)
        .select_related(
            'registration__competitor',
            'registration__competition_day__discipline',
            'registration__competition_day__stage__season',
            'registration__age_group'
        )
    )
    results_list = list(results_qs)

    # 4) Sort them by the front-end's custom order
    #    i.e. if resultIds is [12,55,41], we do 12->55->41
    results_list.sort(key=lambda r: index_map[r.id])

    if not results_list:
        return Response({"message": "No matching results found."}, status=200)

    # 5) Group them => day => age_group => [Results in EXACT front-end order]
    # We'll use OrderedDict to preserve insertion order
    from collections import OrderedDict
    day_group_map = OrderedDict()

    for r in results_list:
        cday = r.registration.competition_day
        ag = r.registration.age_group

        if cday not in day_group_map:
            day_group_map[cday] = OrderedDict()

        if ag not in day_group_map[cday]:
            day_group_map[cday][ag] = []

        day_group_map[cday][ag].append(r)

    # 6) The color scheme
    sky_color = colors.HexColor('#2F89D0')   # For day headers
    sand_color = colors.HexColor('#b0caff')  # For age-group headers
    page_width, page_height = letter
    table_width = page_width * 0.9

    # We'll define columns => [Place, BIB, Name, YOB, School, run1, run2, total? etc.]
    # or we can keep it minimal
    column_widths = [
        table_width / 16,  # Place
        table_width / 16,  # BIB
        table_width * 0.25,  # Name
        table_width / 10,   # YOB
        table_width * 0.20, # School
        table_width / 10,   # run1
        table_width / 10,   # run2
        table_width / 8     # total_time
    ]

    # 7) Build the PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Results.pdf"'
    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=1.5 * 72)
    elements = []

    # local function to draw a custom header
    def draw_header(canvas, doc):
        # Example code to draw a logo
        logo_path = os.path.join(settings.BASE_DIR, 'competition/fonts', 'GSF-Logo.png')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=80, height=80)
            logo.drawOn(canvas, doc.leftMargin - 40, doc.height + doc.topMargin - 30)

        title_style = ParagraphStyle(name='TitleStyle', fontSize=16, alignment=1)
        p_title = Paragraph("Results PDF", title_style)
        p_title.wrapOn(canvas, doc.width, doc.topMargin)
        p_title.drawOn(canvas, doc.width / 2 - 60, doc.height + doc.topMargin)

        date_text = "Date: " + datetime.now().strftime('%Y-%m-%d')
        p_date = Paragraph(date_text, title_style)
        p_date.wrapOn(canvas, doc.width, doc.topMargin)
        p_date.drawOn(canvas, doc.width - doc.rightMargin - 110, doc.height + doc.topMargin)

    # 8) AgeGroup sort function: female -> male => descending birth_year_start
    def ag_sort_key(ag):
        """
        1) If ag.gender == 'ქალი', gender_priority = 0
           If ag.gender == 'კაცი', gender_priority = 1
        2) birth_year_start = ag.birth_year_start or 0
           Negative => descending
        """
        gender_priority = 0 if ag.gender.name == 'ქალი' else 1
        start_val = ag.birth_year_start or 0
        return (gender_priority, -start_val)

    # Build content
    for cday, ag_map in day_group_map.items():
        # cday header
        day_header_data = [[f"{cday.discipline.name} on {cday.date} | {cday.stage.name} - {cday.stage.season.season}"]]
        day_header_table = Table(day_header_data, colWidths=[table_width])
        day_header_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), sky_color),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONT', (0, 0), (-1, -1), 'kartuli', 12),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ])
        day_header_table.setStyle(day_header_style)
        elements.append(day_header_table)
        elements.append(Spacer(1, 12))

        # Sort the age groups
        sorted_ags = sorted(ag_map.keys(), key=ag_sort_key)

        for ag in sorted_ags:
            results_in_ag = ag_map[ag]

            # build age group string, e.g. "ქალი 2010-2011"
            if ag.birth_year_start is None and ag.birth_year_end is None:
                year_range = "None"
            elif ag.birth_year_start is None:
                year_range = f"↓ - {ag.birth_year_end}"
            elif ag.birth_year_end is None:
                year_range = f"{ag.birth_year_start}+"
            else:
                year_range = f"{ag.birth_year_start}-{ag.birth_year_end}"

            ag_header_data = [[f"{ag.gender} {year_range}"]]
            ag_header_table = Table(ag_header_data, colWidths=[table_width])
            ag_header_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), sand_color),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONT', (0, 0), (-1, -1), 'kartuli', 10),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ])
            ag_header_table.setStyle(ag_header_style)
            elements.append(ag_header_table)
            elements.append(Spacer(1, 8))

            # Build table data
            # We'll do: [Place, BIB, Name, YOB, School, run1, run2, total_time]
            data = [[
                'Place', 'BIB', 'Name', 'Year', 'School',
                'Run1', 'Run2', 'Total'
            ]]
            for r in results_in_ag:
                comp = r.registration.competitor
                data.append([
                    r.place or '',
                    r.registration.bib_number or '',
                    f"{comp.first_name} {comp.last_name}",
                    comp.year_of_birth or '',
                    comp.school or '',
                    r.run1_time or '',
                    r.run2_time or '',
                    r.total_time or ''
                ])

            # Table
            data_table = Table(data, colWidths=column_widths)
            data_style = TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONT', (0, 0), (-1, -1), 'kartuli', 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
            ])
            data_table.setStyle(data_style)
            elements.append(data_table)
            elements.append(Spacer(1, 12))

    # Build PDF
    doc.build(elements, onFirstPage=draw_header)
    return response



@api_view(['GET'])
@permission_classes([AllowAny])  # or IsAuthenticated if you prefer
def season_winners(request):
    """
    GET /api/season-winners/?season=2023-2024

    Summarizes each competitor's total season_points across all competition days
    in the given season, grouped by age group. Tie-break = sum of places ascending.
    If place is None => we treat it as 999 (a big but not insane sentinel).
    """
    season_name = request.GET.get('season')
    if not season_name:
        return Response({"error": "Missing ?season=XXXX-YYYY param"}, status=400)

    # 1) Find the Season object
    try:
        season_obj = Season.objects.get(season=season_name)
    except Season.DoesNotExist:
        return Response({"error": f"No Season named {season_name} found"}, status=404)

    # 2) Fetch all Results for that season
    season_results = (
        Result.objects
        .select_related(
            'registration__competitor',
            'registration__age_group',
            'registration__competition_day__stage__season'
        )
        .filter(registration__competition_day__stage__season=season_obj)
    )
    if not season_results.exists():
        return Response({
            "season": season_name,
            "age_groups": [],
            "message": "No results found for this season"
        }, status=200)

    # 3) Build a dict: { ageGroup => { competitor_id => { sum_points, sum_places } } }
    from collections import defaultdict
    ag_competitors = defaultdict(lambda: defaultdict(lambda: {"sum_points": 0, "sum_places": 0}))

    for r in season_results:
        ag = r.registration.age_group
        competitor = r.registration.competitor
        cid = competitor.id

        # Sum up the competitor's season_points
        # place => if None => 999 (so DNF => big place)
        sp = r.points or 0
        place = r.place if r.place is not None else 999

        ag_competitors[ag][cid]["sum_points"] += sp
        ag_competitors[ag][cid]["sum_places"] += place

    # 4) We'll sort AgeGroups female->male => youngest->oldest
    def ag_sort_key(ag):
        # If your 'ag.gender' is a string, do:
        gender_priority = 0 if ag.gender.name == 'ქალი' else 1
        start_val = ag.birth_year_start or 0
        return (gender_priority, -start_val)  # descending

    # Sort the AgeGroups
    sorted_ags = sorted(ag_competitors.keys(), key=ag_sort_key)

    # 5) Build the final JSON
    final_result = {
        "season": season_name,
        "age_groups": []
    }

    # For each age group, we produce a sorted list of winners
    # Sort => season_points desc, sum_places asc
    for ag in sorted_ags:
        competitor_map = ag_competitors[ag]  # => { cid => { sum_points, sum_places } }
        comp_list = []
        for cid, sums in competitor_map.items():
            comp_list.append((cid, sums["sum_points"], sums["sum_places"]))

        # Sort => highest sum_points first, tie => lowest sum_places
        comp_list.sort(key=lambda x: (-x[1], x[2]))

        # Build JSON sublist
        winners = []
        rank = 1
        # We'll fetch competitor data from the DB
        from .models import Competitor  # or wherever your competitor model is
        for (cid, sum_p, sum_pl) in comp_list:
            # get competitor's name
            competitor_obj = Competitor.objects.get(pk=cid)
            winners.append({
                "competitor_id": cid,
                "first_name": competitor_obj.first_name,
                "last_name": competitor_obj.last_name,
                "season_points": sum_p,
                "sum_places": sum_pl,
                "ranking": rank
            })
            rank += 1

        final_result["age_groups"].append({
            "gender": ag.gender.name if ag.gender else "",
            "birth_year_start": ag.birth_year_start,
            "birth_year_end": ag.birth_year_end,
            "winners": winners
        })

    return Response(final_result, status=200)