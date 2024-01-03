from django.shortcuts import render,redirect
from .models import *
from .serializers import *
from rest_framework import generics
from rest_framework.decorators import api_view,permission_classes
import random
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework.authtoken.models import Token
from rest_framework.views import APIView
from rest_framework.response import Response
from django.contrib.auth import authenticate
from rest_framework import status
from rest_framework.authentication import TokenAuthentication
from django.db import IntegrityError
import pandas as pd
from django.http import HttpResponse
from django.db.models import Prefetch
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ObtainAuthTokenView(APIView):
    throttle_classes = ()
    permission_classes = ()
    authentication_classes = ()

    @swagger_auto_schema(
        operation_description="Post username and password to get an authentication token",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['username', 'password'],
            properties={
                'username': openapi.Schema(type=openapi.TYPE_STRING, description='Username'),
                'password': openapi.Schema(type=openapi.TYPE_STRING, description='Password'),
            },
        ),
        responses={
            200: TokenSerializer,
            400: 'Invalid username/password'
        }
    )
    def post(self, request, *args, **kwargs):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            token, created = Token.objects.get_or_create(user=user)
            return Response({'token': token.key})
        else:
            return Response({"error": "Wrong Credentials"}, status=status.HTTP_400_BAD_REQUEST)



class LogoutView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_description="Logout a user by invalidating the authentication token",
        responses={
            200: 'Successfully logged out',
            401: 'Unauthorized'
        }
    )
    def post(self, request, *args, **kwargs):
        # This will only succeed if the user is authenticated and has a valid token
        request.user.auth_token.delete()
        return Response({"message": "Successfully logged out"}, status=status.HTTP_200_OK)




class SchoolListCreate(generics.ListCreateAPIView):
    queryset = School.objects.all()
    serializer_class = SchoolSerializer
    # Set different permissions for different request methods
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class SchoolRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = School.objects.all()
    serializer_class = SchoolSerializer
    permission_classes = [IsAuthenticated]


class DisciplineListCreate(generics.ListCreateAPIView):
    queryset = Discipline.objects.all()
    serializer_class = DisciplineSerializer
    # Set different permissions for different request methods
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class DisciplineRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Discipline.objects.all()
    serializer_class = DisciplineSerializer
    permission_classes = [IsAuthenticated]


class CompetitorListCreate(generics.ListCreateAPIView):
    queryset = Competitor.objects.all()
    serializer_class = CompetitorSerializer
    # Set different permissions for different request methods
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]




class CompetitorRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Competitor.objects.all()
    serializer_class = CompetitorSerializer
    permission_classes = [IsAuthenticated]



class SeasonListCreate(generics.ListCreateAPIView):
    queryset = Season.objects.all()
    serializer_class = SeasonSerializer
    # Set different permissions for different request methods
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class SeasonRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Season.objects.all()
    serializer_class = SeasonSerializer
    permission_classes = [IsAuthenticated]


class CompetitionDayCreate(generics.ListCreateAPIView):
    queryset = CompetitionDay.objects.all()
    serializer_class = CompetitionDaySerializer
    # Set different permissions for different request methods
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class CompetitionDayRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = CompetitionDay.objects.all()
    serializer_class = CompetitionDaySerializer
    permission_classes = [IsAuthenticated]



class StageListCreate(generics.ListCreateAPIView):
    queryset = Stage.objects.all()
    serializer_class = StageSerializer
    # Set different permissions for different request methods
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class StageRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Stage.objects.all()
    serializer_class = StageSerializer
    permission_classes = [IsAuthenticated]




class GroupListCreate(generics.ListCreateAPIView):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    # Set different permissions for different request methods
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class GroupRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = [IsAuthenticated]



class CartListCreate(generics.ListCreateAPIView):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer
    # Set different permissions for different request methods
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]


class CartRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer
    permission_classes = [IsAuthenticated]

class ResultsListCreate(generics.ListCreateAPIView):
    queryset = Results.objects.all()
    serializer_class = ResultsSerializer
    # Set different permissions for different request methods
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

class ResultsRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Results.objects.all()
    serializer_class = ResultsSerializer
    permission_classes = [IsAuthenticated]


@swagger_auto_schema(method='post', request_body=RandomizeBibNumbersSerializer)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def randomize_bib_numbers(request):
    serializer = RandomizeBibNumbersSerializer(data=request.data)
    if serializer.is_valid():
        start_number = serializer.validated_data.get('start_number')
        ignore_numbers = serializer.validated_data.get('ignore_numbers', [])
        competition_id = serializer.validated_data.get('competition', None)
        gender_id = serializer.validated_data.get('gender', None)

        # Fetch all groups for the specified competition and gender
        groups = Group.objects.filter(competition_id=competition_id, gender_id=gender_id)

        # Loop through each group to randomize bib numbers
        for group in groups:
            # Fetch carts for the current group
            carts = Cart.objects.filter(group=group)
            # Calculate the total number of carts in the group
            total_carts = carts.count()

            # Generate potential numbers for the current group
            potential_numbers = set(range(start_number, start_number + total_carts))

            # Check if there is an overlap between potential numbers and ignored numbers
            if any(num in potential_numbers for num in ignore_numbers):
                # Extend the potential range until there is no overlap
                potential_numbers = set(range(start_number, start_number + total_carts + len(ignore_numbers)))

            
            try:
                start_number = max(potential_numbers) + 1
            except:
                print("no more groups")
            # Remove ignored numbers from potential numbers
            potential_numbers -= set(ignore_numbers)

            # Check if there are enough unique numbers available
            if len(potential_numbers) < total_carts:
                return Response({"error": "Not enough unique numbers available after excluding the ignored numbers."}, status=400)

            # Randomize bib numbers for carts in the current group
            for cart in carts:
                cart.bib_number = random.choice(list(potential_numbers))
                potential_numbers.remove(cart.bib_number)
                cart.save()

            # Update start_number for the next grou

        return Response({"message": "Bib numbers randomized successfully"})
    else:
        return Response(serializer.errors, status=400)








def create_result_for_cart(cart_id):
    default_status, created = Status.objects.get_or_create(status='Active')

    try:
        cart = Cart.objects.get(pk=cart_id)
        result_exists = Results.objects.filter(competitor=cart).exists()

        if not result_exists:
            Results.objects.create(
                competitor=cart
            )
        else:
            # If a result already exists, you may want to handle this case accordingly
            pass

    except Cart.DoesNotExist:
        return Response({"error": f"Cart with ID {cart_id} does not exist"}, status=status.HTTP_404_NOT_FOUND)
    except IntegrityError:
        return Response({"error": f"Error creating result for Cart ID {cart_id}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response({"message": f"Result added successfully for Cart ID {cart_id}"}, status=status.HTTP_200_OK)







@api_view(['POST'])
@permission_classes([IsAuthenticated])
def batch_sync_results(request):
    cart_ids = request.data.get('cart_ids', [])

    # Iterate over cart IDs and create results
    for cart_id in cart_ids:
        create_result_for_cart(cart_id)

    return Response({"message": "Results added successfully"})




class SyncCartToResultsView(APIView):
    permission_classes = [IsAuthenticated]  # Ensure only admin users can access this endpoint

    @swagger_auto_schema(
        operation_description="Synchronize Cart entries to Results table",
        responses={
            200: 'Synchronization completed successfully',
            500: 'Internal Server Error'
        }
    )
    def post(self, request, *args, **kwargs):
        try:
            # Perform the synchronization
            sync_cart_to_results()
            return Response({"message": "Synchronization completed successfully"}, status=status.HTTP_200_OK)
        except Exception as e:
            # Log the error or handle it as needed
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def sync_cart_to_results():
    default_status, created = Status.objects.get_or_create(status='Active')

    for cart in Cart.objects.all():
        if cart.competitor:
            result_exists = Results.objects.filter(competitor=cart).exists()

            if not result_exists:
                try:
                    Results.objects.create(
                        competitor=cart,
                        status=default_status
                    )
                except IntegrityError as e:
                    # Handle the exception or log it
                    pass  # Replace with actual error handling

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def download_excel(request):
    if request.method == 'POST':
        cart_ids = request.data.get('cartIds', [])

        groups_with_carts = Group.objects.prefetch_related(
            Prefetch('cart_set', queryset=Cart.objects.filter(id__in=cart_ids).order_by('bib_number'), to_attr='sorted_carts')
        ).filter(cart__id__in=cart_ids).distinct().order_by('id')

        # List to hold all the DataFrames
        dfs = []
        bold_font = Font(bold=True)  # Create a Font object with bold text
        sky_blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')  # Sky blue fill

        for group in groups_with_carts:
            # Add a row with the group name
            group_name_row = pd.DataFrame([{'BIB': 'Group:', 'Name': group.group_name}])
            dfs.append(group_name_row)

            # Add the data for each group
            group_data = [{
                'BIB': cart.bib_number,
                'Name': f'{cart.competitor.name} {cart.competitor.surname}',
                'Gender': cart.competitor.gender,
                'Year': cart.competitor.year,
                'School': cart.competitor.school.school_name
            } for cart in group.sorted_carts]

            group_df = pd.DataFrame(group_data)
            dfs.append(group_df)

        # Concatenate all DataFrames
        all_data = pd.concat(dfs, ignore_index=True)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Competitors.xlsx"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            all_data.to_excel(writer, sheet_name='Competitors', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Competitors']

            # Apply bold font and sky blue background to group name rows
            for row in worksheet.iter_rows(min_row=1, max_col=2, max_row=worksheet.max_row):
                for cell in row:
                    if cell.value == 'Group:':
                        for cell in row:
                            cell.font = bold_font
                            cell.fill = sky_blue_fill

        return response
    return Response({"error": "Invalid request"}, status=400)


